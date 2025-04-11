from datetime import datetime
from datetime import timedelta
import numbers
import os
import shutil
import traceback

import openpyxl
from openpyxl.styles import Font  # Import Font from styles
from openpyxl.formatting.rule import CellIsRule
from openpyxl.utils import get_column_letter
import math

#import settings
import re

import openpyxl.utils




def get_input_files():
    """
    Gets all .xlsx files from a subfolder 'excel' located in the same directory as the script.
    Returns a list of file paths.
    """
    # Get the current script's directory
    base_dir = os.path.dirname(os.path.abspath(__file__))
    base_dir += "//VISTA_INPUT_FOLDER"
    print(base_dir)

    # Target 'excel' folder within that directory
    #target_dir = os.path.join(base_dir, "excel")

    input_files = []
    for entity in os.listdir(base_dir):
        path = os.path.join(base_dir, entity)
        # Check if it's a file and ends with .xlsx (case-insensitive)
        if os.path.isfile(path) and entity.lower().endswith('.xlsx'):
            input_files.append(path)

    return input_files


def generate_title(file_path):
    """
        Generates a new title with the actual date and time attached
        Returns the new filename.
    """
    #date = datetime.strftime(datetime.now(), '%m-%d-%y %X')
    new_filename = os.path.basename(file_path)

    return new_filename


def move_to_output(file_path):
    """
        Moves a file to Output Folder
    """
    # Get the current script's directory
    base_dir = os.path.dirname(os.path.abspath(__file__))
    base_dir += "//VISTA_OUTPUT_FOLDER"

    new_filename = generate_title(file_path)
    try:
        os.remove(os.path.join(base_dir, new_filename))
    except:
        pass
    shutil.move(file_path, os.path.join(base_dir, new_filename))

def move_to_output_failed(file_path):
    """
        Moves a file to Output Folder
    """
    # Get the current script's directory
    base_dir = os.path.dirname(os.path.abspath(__file__))
    base_dir += "//VISTA_FAILED_OUTPUT_FOLDER"

    new_filename = generate_title(file_path)
    try:
        os.remove(os.path.join(base_dir, new_filename))
    except:
        pass
    shutil.move(file_path, os.path.join(base_dir, new_filename))


def open_excel_book(filename, data_only=False):
    return openpyxl.load_workbook(filename, data_only=data_only)


def label_sheet(worksheet):
    label = worksheet.title
    worksheet['A2'] = label


def add_totals(worksheet, row, column, msg='Grand Totals: '):
    first_row = 6
    min_column = 7 if worksheet.title != 'Grouped Jobs' else 8

    worksheet.cell(row + 2, min_column - 1).value = msg

    for col in range(min_column + 1, column):
        # print(openpyxl.utils.get_column_letter(col))
        if worksheet.title != 'Grouped Jobs' and openpyxl.utils.get_column_letter(col) in ['V', 'W', 'X', 'Y', 'Z', 'AA', 'AB', 'AC', 'AE', 'AL']:
            worksheet[
                f'{openpyxl.utils.get_column_letter(col)}{row + 2}'] = f'=AVERAGE({openpyxl.utils.get_column_letter(col)}{first_row}:{openpyxl.utils.get_column_letter(col)}{row})'
        elif worksheet.title == 'Grouped Jobs' and openpyxl.utils.get_column_letter(col) in ['W', 'X', 'Y', 'Z', 'AA', 'AB', 'AC', 'AD', 'AF', 'AM']:
            worksheet[
                f'{openpyxl.utils.get_column_letter(col)}{row + 2}'] = f'=AVERAGE({openpyxl.utils.get_column_letter(col)}{first_row}:{openpyxl.utils.get_column_letter(col)}{row})'
        else:
            worksheet[
                f'{openpyxl.utils.get_column_letter(col)}{row + 2}'] = f'=SUM({openpyxl.utils.get_column_letter(col)}{first_row}:{openpyxl.utils.get_column_letter(col)}{row})'

        if openpyxl.utils.get_column_letter(col) in ['V', 'W', 'X', 'Y', 'Z', 'AA', 'AB', 'AC', 'AE', 'AL']:
            worksheet[f'{openpyxl.utils.get_column_letter(col)}{row + 2}'].number_format = '0.00%'
        else:
            worksheet[f'{openpyxl.utils.get_column_letter(col)}{row + 2}'].number_format = '#,##0.00'

def add_departments(worksheet, row):
    first_row = 6
    min_column = 7 if worksheet.title != 'Grouped Jobs' else 8
    department_string_values = []
    department_cost_values = {}

    worksheet.cell(row + 5, min_column - 1).value = "Department Totals"
    worksheet.cell(row + 5, min_column + 1).value = "Billings In Excess Of Costs And Estimated Earnings"

    for i in range(7, row + 1):
        department_string = worksheet.cell(i, min_column - 1).value
        department_value = worksheet.cell(i, 42).value
        if department_string not in department_string_values:
            worksheet.cell(row + 6 + len(department_string_values), min_column).value = department_string
            department_string_values.append(department_string)
            department_cost_values[department_string] = []
            department_cost_values[department_string].append(i)
        else:            
            department_cost_values[department_string].append(i)

    for i in range(row + 6,  row + 6 + len(department_string_values)):
        # Get column letter
        col_letter = openpyxl.utils.get_column_letter(42)

        key_str = list(department_cost_values.values())[i - row - 6]        
        # Build formula parts in each departments
        cell_refs = [f"{col_letter}{row}" for row in key_str]

        # Join them into a SUM formula
        sum_formula = f"=SUM({', '.join(cell_refs)})"

        # Assign to cell
        worksheet[
                f'{openpyxl.utils.get_column_letter(min_column + 1)}{i}'] = sum_formula
        
        worksheet.cell(i, min_column + 1).number_format = worksheet.cell(8, 42).number_format




def format_number(value):
    # Remove negative sign
    value = abs(value)

    # Cut (not round) to 2 decimal places
    value = math.floor(value * 100) / 100

    # Format with commas and 2 decimal places
    return value

def get_last_coord(worksheet):
    last_row = worksheet.max_row
    last_col = worksheet.max_column

    while worksheet[f'A{last_row}'].value is None:
        last_row -= 1

    while worksheet[f'{openpyxl.utils.get_column_letter(last_col)}6'].value is None:
        last_col -= 1

    return last_row, last_col


def apply_negative_red_formatting(worksheet, cell_range):

    red_font = Font(color="9C0006")  # Un tono de rojo
    worksheet.conditional_formatting.add(
        cell_range,
        CellIsRule(operator='lessThan', formula=['0'], font=red_font)
    )


def add_subtotals(worksheet, group_list: list, row_number: int, column, offset):
    min_column = 7

    worksheet.cell(row_number, min_column).value = 'Subtotal: '

    for col in range(min_column + 1, column):
        if openpyxl.utils.get_column_letter(col) in ['W', 'X', 'Y', 'Z', 'AA', 'AB', 'AC', 'AD', 'AF', 'AM']:
            worksheet[
                f'{openpyxl.utils.get_column_letter(col)}{row_number}'] = f'=AVERAGE({openpyxl.utils.get_column_letter(col)}{group_list[0] + offset}:{openpyxl.utils.get_column_letter(col)}{group_list[-1] + offset})'
        else:
            worksheet[
                f'{openpyxl.utils.get_column_letter(col)}{row_number}'] = f'=SUM({openpyxl.utils.get_column_letter(col)}{group_list[0] + offset}:{openpyxl.utils.get_column_letter(col)}{group_list[-1] + offset})'

        if openpyxl.utils.get_column_letter(col) in ['W', 'X', 'Y', 'Z', 'AA', 'AB', 'AC', 'AD', 'AF', 'AM']:
            worksheet[f'{openpyxl.utils.get_column_letter(col)}{row_number}'].number_format = '0.00%'
        else:
            worksheet[f'{openpyxl.utils.get_column_letter(col)}{row_number}'].number_format = '#,##0.00'


def format_subtotals(worksheet):
    groups = {}
    offset = 0
    rows, cols = get_last_coord(worksheet)

    for row in range(7, rows + 1):
        if worksheet[f'B{row}'].value not in groups.keys():
            groups[worksheet[f'B{row}'].value] = []
        groups[worksheet[f'B{row}'].value].append(row)

    for group in groups.keys():
        last_group_row = groups[group][-1] + offset
        worksheet.insert_rows(last_group_row + 1)
        add_subtotals(worksheet, groups[group], last_group_row + 1, cols, offset)
        worksheet.insert_rows(last_group_row + 2)
        groups[group] = [row + offset for row in groups[group]]
        offset += 2

    return groups


def remove_rows(worksheet, last_row):
    worksheet.delete_rows(last_row+1, 50)


def update_totals(worksheet, last_row, last_column, grouped_rows: dict):
    min_column = 8
    row_intervals = ''

    for group in grouped_rows.values():
        row_intervals += f'_{group[0]}:_{group[-1]},'
    row_intervals = row_intervals[:-1]

    worksheet.cell(last_row + 4, min_column).value = 'Grand Totals: '

    for col in range(min_column + 1, last_column):
        if openpyxl.utils.get_column_letter(col) in ['W', 'X', 'Y', 'Z', 'AA', 'AB', 'AC', 'AD', 'AF', 'AM']:
            worksheet[f'{openpyxl.utils.get_column_letter(col)}{last_row + 4}'] = u'=AVERAGE({})'.format(
                row_intervals.replace("_", openpyxl.utils.get_column_letter(col)))
        else:
            worksheet[f'{openpyxl.utils.get_column_letter(col)}{last_row + 4}'] = u'=SUM({})'.format(
                row_intervals.replace("_", openpyxl.utils.get_column_letter(col)))

        if openpyxl.utils.get_column_letter(col) in ['W', 'X', 'Y', 'Z', 'AA', 'AB', 'AC', 'AD', 'AF', 'AM']:
            worksheet[f'{openpyxl.utils.get_column_letter(col)}{last_row + 4}'].number_format = '0.00%'
        else:
            worksheet[f'{openpyxl.utils.get_column_letter(col)}{last_row + 4}'].number_format = '#,##0.00'


def extend_table_limits(worksheet, last_row, last_column):
    table = list(worksheet.tables.values())[0]
    # print(table.ref)
    table.ref = f'A6:{openpyxl.utils.get_column_letter(last_column)}{last_row + 4}'


def get_jobs_numbers(excel_book):
    jobs = []
    for worksheet in excel_book:
        last_row, last_column = get_last_coord(worksheet)
        if worksheet.title in ['Contracts > $1MIL', 'Contracts < $1MIL', 'Completed Contracts > $1MIL',
                               'Completed Contracts < $1MIL']:
            for row in worksheet.iter_rows(min_row=7, max_row=last_row):
                jobs.append(row[0].value)

    return jobs


def remove_repeated_jobs_in_grouped_sheet(excel_book):
    jobs = []
    worksheet = excel_book['Grouped Jobs']
    last_row, last_column = get_last_coord(worksheet)
    for row in range(last_row + 1, 6, -1):
        if worksheet[row][0].value not in jobs and worksheet[row][0].value is not None:
            jobs.append(worksheet[row][0].value)

    for worksheet in excel_book:
        last_row, last_column = get_last_coord(worksheet)
        if worksheet.title in ['Contracts > $1MIL', 'Contracts < $1MIL', 'Completed Contracts > $1MIL',
                               'Completed Contracts < $1MIL']:
            for row in range(last_row + 1, 6, -1):
                if worksheet[row][0].value in jobs:
                    worksheet.delete_rows(idx=row, amount=1)


def grouped_jobs_wih_info(excel_book, jobs):
    worksheet = excel_book['Grouped Jobs']
    last_row, last_column = get_last_coord(worksheet)

    grouped_jobs = {}
    # get column headers
    headers = []
    for index, col in enumerate(worksheet.iter_cols(min_row=6, max_row=6, max_col=last_column)):
        headers.append(col[0].value)

    # iterate over rows
    for row in worksheet.iter_rows(min_row=7, max_row=last_row):
        job = row[0].value
        group_name = row[1].value
        # if group_name exists
        if job not in jobs and group_name:
            if not grouped_jobs.get(group_name, None):
                grouped_jobs[group_name] = {}
            for col in range(0, len(headers)):
                # for save None values
                if grouped_jobs[group_name].get(headers[col], False) is False:
                    if row[col].data_type == 'f':
                        grouped_jobs[group_name][headers[col]] = re.sub(r'Table\d+\[', '[', row[col].value)
                    else:
                        grouped_jobs[group_name][headers[col]] = row[col].value
                elif isinstance(grouped_jobs[group_name][headers[col]], numbers.Number):
                    grouped_jobs[group_name][headers[col]] += row[col].value

    # remove contracts
    for k, v in grouped_jobs.items():
        v[headers[0]] = 'Various'
        v.pop(headers[2])

    return [x for x in grouped_jobs.values()]


def insert_grouped_jobs_into_contracts(excel_book, grouped_jobs):
    if grouped_jobs:
        headers = [x for x in grouped_jobs[0].keys()]
        worksheet = excel_book['Contracts > $1MIL']
        last_row, last_column = get_last_coord(worksheet)
        # extend_table_limits(worksheet, last_row + len(grouped_jobs) + 10, last_column)
        worksheet.insert_rows(idx=7, amount=len(grouped_jobs))

        for row_index, row in enumerate(worksheet.iter_rows(min_row=7, max_row=len(grouped_jobs) + 6)):
            for col in range(0, len(headers)):
                row[col].value = grouped_jobs[row_index][headers[col]]
                if isinstance(row[col].value, numbers.Number) or row[col].data_type == 'f' or col == len(headers) - 1:
                    if '%' in headers[col]:
                        row[col].number_format = '0.00%'
                    elif col == len(headers) - 1:
                        row[col].number_format = 'MM/DD/YYYY'
                    else:
                        row[col].number_format = '#,##0.00'


def add_grouped_totals(excel_book):
    rows = {}
    for worksheet in excel_book:
        if worksheet.title in ['Contracts < $1MIL', 'Completed Contracts > $1MIL', 'Completed Contracts < $1MIL']:
            # calculate last row
            last_row = worksheet.max_row
            last_col = worksheet.max_column - 1
            while worksheet[f'G{last_row}'].value is None:
                last_row -= 1

            # initialize dict
            rows[worksheet.title] = []
            if worksheet.title in ['Contracts < $1MIL', 'Completed Contracts > $1MIL', 'Completed Contracts < $1MIL']:
                for col in worksheet.iter_cols(min_row=last_row, max_row=last_row, min_col=8, max_col=last_col):
                    rows[worksheet.title].append(f"='{worksheet.title}'!{col[0].coordinate}")
    return rows


def write_grouped_totals(excel_book, grouped_totals):
    worksheet = excel_book['Contracts > $1MIL']
    last_row, last_column = get_last_coord(worksheet)

    for k, v in grouped_totals.items():
        worksheet.cell(row=last_row + 9, column=1).value = k
        for index, cell in enumerate(v, start=1):
            worksheet.cell(row=last_row + 9, column=index + 7).value = cell
            if openpyxl.utils.get_column_letter(index + 5) in ['V', 'W', 'X', 'Y', 'Z', 'AA', 'AB', 'AC', 'AE', 'AL']:
                worksheet.cell(row=last_row + 9, column=index + 7).number_format = '0.00%'
            else:
                worksheet.cell(row=last_row + 9, column=index + 7).number_format = '#,##0.00'
        last_row += 1


def update_column_headers(worksheet, file_date):
    """
    Updates the headers in the worksheet to replace 'Month1', 'Month2', 'Month3'
    with the correct month names based on the file's date.
    """
    # Calculate previous three months based on the file_date
    month1 = (file_date - timedelta(days=31)).strftime('%b %y')  # Jun 24
    month2 = (file_date - timedelta(days=61)).strftime('%b %y')  # May 24
    month3 = (file_date - timedelta(days=92)).strftime('%b %y')  # Apr 24

    # Define the column names to be changed
    headers_map = {
        "Month1": month1,
        "Month2": month2,
        "Month3": month3
    }

    last_row, last_column = get_last_coord(worksheet)

    # Iterate through the worksheet headers and replace as needed
    for row in worksheet.iter_rows(min_row=6, max_row=6, min_col=1, max_col=last_column):
        for cell in row:
            if cell.value:
                for key, value in headers_map.items():
                    if key in cell.value:
                        cell.value = cell.value.replace(key, value)

def main():
    files = get_input_files()
    for file in files:
        try:

            #open the each one excel file by uisng directory
            excel_book = open_excel_book(file)

            #calculate total values and Department values in each worksheet in each one excel file
            for worksheet in excel_book:
                label_sheet(worksheet)
                last_row, last_column = get_last_coord(worksheet)
                print(worksheet.title, last_row, last_column)

                remove_rows(worksheet, last_row)

                if worksheet.title != 'Contracts > $1MIL':
                    add_totals(worksheet, last_row, last_column)
                    add_departments(worksheet, last_row)
            
            #save excel file
            excel_book.save(file)
            excel_book.close()
            move_to_output(file)

        except Exception as e:
            traceback.print_exc()
            print(e)
            move_to_output_failed(file)

    # Delete every file in c:/Temp folder
    try:
        for file in os.listdir('C:/Temp'):
            if os.path.isfile(os.path.join('C:/Temp', file)):
                os.remove(os.path.join('C:/Temp', file))
    except:
        pass


if __name__ == '__main__':
    main()
